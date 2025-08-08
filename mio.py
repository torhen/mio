# v2.1
USE_GEOPANDAS = True

import sys,os,datetime
import pandas as pd
import numpy as np
import subprocess, time, datetime
from PIL import Image
import pathlib
import sqlite3
import json
import xlwings as xw
import openpyxl
from copy import copy
import sqlalchemy
pd.options.display.max_columns = 0

# --------- Database connections ----------
def read_sql(sql, database, make_upper=True):

    if database.upper() == 'NETSITE':  
        con = sqlalchemy.create_engine('oracle://bo_readonly:reporter@pnsite01')
    elif database.upper() == 'DPI':
        con = sqlalchemy.create_engine('oracle://DPIREAD:DPIREAD@p1dpich')
    elif database.upper() == 'ATOLL':
        con = sqlalchemy.create_engine('oracle://COVERAGE_MAPS:COST!231%Walfisch$Ikegami@p1forch')
    elif database.upper() == 'MCMS' or database.upper() == 'PMD':
        con = sqlalchemy.create_engine('oracle://PMDREAD:PMDREAD@ppmd01')
    elif database.upper() == 'SIP':
        # con_ip = cx_Oracle.connect('READONLY', 'R35d0nly', 'PSIP01')
        con = sqlalchemy.create_engine('oracle://READONLY:R35d0nly@PSIP01')
    elif database.upper() == 'DWH' or database.upper() == 'TERA' or database.upper() == 'TERADATA':
        #pip install teradatasqlalchemy
        con = sqlalchemy.create_engine(f'teradatasql://FnwserverA:Fnwserver%40Teradata20%218@P1EDWCH') # @ in the password coded!
    else:
        assert False, f"Dont't know database {database}"

    data = pd.read_sql(sql, con)
    
    if make_upper:
            data = data.rename(columns=str.upper)
    return data



def join_ordered(l):
    l = [str(s) for s in l if pd.notnull(s)]
    l = list(set(l))
    l.sort()
    return ';'.join(l)

def now():
    return str(datetime.datetime.now())[0:19]

if USE_GEOPANDAS:
    import geopandas as gpd
    from shapely.geometry import Point, Polygon, MultiPolygon, LineString, MultiLineString, shape
    import rasterio
    import rasterio.features
    import rasterio.mask
    import fiona
    import shapely
    
import matplotlib.pyplot as plt
import nbformat
from nbconvert.preprocessors import ExecutePreprocessor
import win32com.client
import affine

def console(text):
    ts = str(datetime.datetime.now())[0:19]
    b = bytes(f'>> {ts} {text}\n', encoding='latin-1')
    os.write(1, b)

def get_time():
    print(str(datetime.datetime.now())[0:19])

def colset(df, cols_dic):
    """ filter and rename columns in one step"""
    return df[list(cols_dic)].rename(columns=cols_dic)


# cou can get the WKT string unsing ogrinfor file layer
WKT_SWISS="""PROJCS["unnamed",
GEOGCS["unnamed",
    DATUM["Switzerland_CH_1903",
        SPHEROID["Bessel 1841",6377397.155,299.1528128],
        TOWGS84[660.077,13.551,369.344,-0.804816,-0.577692,-0.952236,5.66]],
    PRIMEM["Greenwich",0],
    UNIT["degree",0.0174532925199433]],
PROJECTION["Swiss_Oblique_Cylindrical"],
PARAMETER["latitude_of_center",46.9524055555],
PARAMETER["central_meridian",7.4395833333],
PARAMETER["false_easting",600000],
PARAMETER["false_northing",200000],
UNIT["Meter",1.0]]
"""

WKT_SWISS_95="""\
PROJCS["unnamed",
    GEOGCS["unnamed",
        DATUM["MIF 999,10,674.374,15.156,405.346",
            SPHEROID["Bessel 1841",6377397.155,299.1528128],
            TOWGS84[674.374,15.156,405.346,0,0,0,0]],
        PRIMEM["Greenwich",0],
        UNIT["degree",0.0174532925199433]],
    PROJECTION["Swiss_Oblique_Cylindrical"],
    PARAMETER["latitude_of_center",46.952405555556],
    PARAMETER["central_meridian",7.439583333333],
    PARAMETER["false_easting",2600000],
    PARAMETER["false_northing",1200000],
    UNIT["Meter",1.0]]
"""

WKT_WGS="""GEOGCS["unnamed",
    DATUM["WGS_1984",
        SPHEROID["WGS 84",6378137,298.257223563],
        TOWGS84[0,0,0,0,0,0,0]],
    PRIMEM["Greenwich",0],
    UNIT["degree",0.0174532925199433]]
"""

MIF_SWISS='CoordSys Earth Projection 25, 1003, "m", 7.4395833333, 46.9524055555, 600000, 200000'
MIF_WGS='CoordSys Earth Projection 1, 104'

def clean_ascii(df):
    df = df.copy()
    def clean(s):
        s = str(s)
        s = [c for c in s if ord(c) < 256]
        return ''.join(s)

    for col in df.columns:
        if col != 'geometry':
            df[col] = df[col].map(clean)
    return df

def check_path(path):
    path = str(path)
    if os.path.isfile(path):
        title = os.path.basename(path)
        bytes = os.path.getsize(path)
        ctime = datetime.datetime.fromtimestamp(os.path.getctime(path))
        ctime = str(ctime)[0:19]
        size, unit = bytes, 'bytes'
        if bytes > 2**10: size, unit = round(bytes / 2**10, 1), 'KB'
        if bytes > 2**20: size, unit = round(bytes / 2**20, 1), 'MB'
        if bytes > 2**30: size, unit = round(bytes / 2**30, 1), 'GB'
        if bytes > 2**40: size, unit = round(bytes / 2**40, 1), 'TB'
        print(f"ok {title}\t({size} {unit})\tcreated: {ctime}")
        return path
    else:
        abs_path = os.path.abspath(path)
        print(f'MISSING FILE: {abs_path}')
        
def check_folders(*folders):
    ret = True
    for folder in folders:
        folder = str(folder)
        if not os.path.isdir(folder):
            abs_path = os.path.abspath(folder)
            print(f'MISSING FOLDER: {abs_path}')
            ret = False
    return ret

def file_title(path:str):
    return os.path.splitext(os.path.basename(path))[0]

def read_dbf(dbfile):
    """read dbase file"""
    dbfile = str(dbfile)
    from simpledbf import Dbf5
    dbf = Dbf5(dbfile)
    pl = dbf.to_dataframe()
    pl.columns = [a.split('\x00')[0] for a in pl.columns] # remove strange characters in columns
    return pl

def run_nb(ju_nb):
    """Execute a jupyter notebook"""

    if len(sys.argv)>2:
        os.environ["JUPYTER_PARAMETER"] = sys.argv[2]
    else:
        os.environ["JUPYTER_PARAMETER"] = ""
    # the jupyter notebook can retrieve os.environ["JUPYTER_PARAMETER"]
    nb = nbformat.read(open(ju_nb), as_version=4)
    ep = ExecutePreprocessor(timeout=600, kernel_name='python3')
    ep.preprocess(nb, {'metadata': {'path': os.path.dirname(ju_nb)}})
    # write sometimes destroys the notebook!
    # better just write a copy
    new_nb_name=os.path.splitext(ju_nb)[0]+'_last_run.ipynb'
    nbformat.write(nb, open(new_nb_name, mode='wt'))
    
def main():
    """to stat notebook from commandline"""
    if len(sys.argv)>1:
        nb = sys.argv[1]
        print('starting',nb)
        run_nb(nb)
    else:
        print('usage: mio.py notebook.jpynb to execute as jupyter notebook')

if __name__ == "__main__":
    main()


def read_raster(raster_file):
    """ Read a raster file and return a list of dataframes"""
    raster_file = str(raster_file) # in case it's a pathlib Path
    ds = rasterio.open(raster_file)
    t = ds.transform
    # changed from t ds.affine
   
    df_list = []
    # band counts is based 1
    for i in range (1,ds.count+1):
        a = ds.read(i)
        df = pd.DataFrame(a)
        
        # set index and columns to world coordinates
        df.columns = [ (t * (x, 0))[0] for x in df.columns]
        df.index =   [ (t * (0 ,y + 1))[1] for y in df.index]
        # y + 1 because of rasterios transformation is for images, not for my dataframe
        # in my datafae I want to keep the lower left corner as index value, not the upper left
        
        df_list.append (df)
    ds.close()
    return df_list

def write_raster(df_list:pd.DataFrame, dest_file, color_map=0):
    dest_file = str(dest_file)
    """ write df raster list to geo tiff together with world file, or Arcview ESRI grid text file
        df  must be 'uint8' to apply color map
        color_map is dictionary like {0:(255,0,0), 1:(0,255,0)}
    """
    driver_dict = {'.tif': 'GTiff', '.txt': 'AAIGrid', '.asc': 'AAIGrid', '.bil': 'EHdr'}
    driver_string = driver_dict[os.path.splitext(dest_file)[1].lower()]
    
    # in case arguent is data frame, not list
    if isinstance(df_list, pd.DataFrame):
        df_list = [df_list]
    
    # create an affine object
    t = calc_affine(df_list[0])

    # get dimension
    bands = len(df_list)
    h, w = df_list[0].shape
    
    # build one 3-dimensional array from the df list
    l = [df.values for df in df_list]
    a = np.array(l)
    dtype = a.dtype

    # write raster, add color_map if defined, TFW = YES: create woldfile!
    with rasterio.open(dest_file, 
                       mode='w', 
                       driver=driver_string, 
                       width=w, height=h, 
                       count=bands, 
                       dtype=dtype, 
                       transform=t, 
                       tfw='YES',
                       crs =WKT_SWISS
                      ) as dst:
        dst.write(a)
        if color_map:
            dst.write_colormap(1, color_map)

def calc_affine(df):
    """generate transorm affine object from raster data frame """
    x0 = df.columns[0]
    y0 = df.index[0]
    dx = df.columns[1] - df.columns[0]
    dy = df.index[1] - df.index[0]
    
    t = affine.Affine(dx, 0, x0 , 0, dy ,y0 - dy) 
    # y0 - dy because anker point is in the south!
    return t

def vectorize(df):
    """ make shapes from raster, genial! """
    t = calc_affine(df)
    a = df.values
    # zeros an nan are left open space, means mask = True!
    maske = (df != 0).fillna(True)
    gdf = gpd.GeoDataFrame(geometry=[])
    geoms  = []
    value = []
    for s,v in rasterio.features.shapes(a,transform=t,mask=maske.values):
        geoms.append(shape(s))
        value.append(v)
    gdf['geometry'] = geoms
    gdf['val']=value
    return gdf

if USE_GEOPANDAS:
    def rasterize(vector_gdf:gpd.GeoDataFrame, raster_df, values_to_burn=128, fill:int=0, all_touched:bool=False):
        """ burn vector features into a raster, input ruster or resolution"""

        # raster_df is integer, create raster with resolution raster_df 
        if isinstance(raster_df, int):
            res = raster_df
            x0, y0, x1, y1 = vector_gdf.geometry.total_bounds
            x0 = int(x0 // res * res - res)
            x1 = int(x1 // res * res + res)
            y0 = int(y0 // res * res - res)
            y1 = int(y1 // res * res + res)
            raster_df = pd.DataFrame(columns=range(x0, x1, res), index=range(y1, y0, -res))

        # no geometry to burn
        if vector_gdf.unary_union.area==0:
            return raster_df
        
        try:
            geom_value_list = zip(vector_gdf.geometry, values_to_burn)
        except:
            geom_value_list = ( (geom, values_to_burn) for geom in vector_gdf.geometry )
        
        t = calc_affine(raster_df)
        
        result = rasterio.features.rasterize(geom_value_list, 
                                             out_shape=raster_df.shape, 
                                             transform=t, 
                                             fill=fill, 
                                             all_touched=all_touched)
        
        res_df = pd.DataFrame(result, columns=raster_df.columns, index=raster_df.index)
        return res_df

def refresh_excel(excel_file):
    """refreshe excel data and pivot tables"""
    excel_file = str(excel_file)
    excel_file = os.path.abspath(excel_file)
    xlapp = win32com.client.DispatchEx("Excel.Application")
    wb = xlapp.workbooks.open(excel_file)
    xlapp.Visible = True
    wb.RefreshAll()
    count = wb.Sheets.Count
    for i in range(count):
        ws = wb.Worksheets[i]
        pivotCount = ws.PivotTables().Count
        for j in range(1, pivotCount+1):
            ws.PivotTables(j).PivotCache().Refresh()
    xlapp.CalculateUntilAsyncQueriesDone()
    xlapp.DisplayAlerts = False
    wb.Save()
    xlapp.Quit()
    
def delete_mapinfo_files(tab_name):
    base = os.path.splitext(tab_name)[0]
    for ext in ['tab', 'dat', 'id', 'map']:
        path = f"{base}.{ext}"
        if os.path.exists(path):
            os.remove(path)

if USE_GEOPANDAS:
    def write_tab(gdf, tab_name, crs_wkt=WKT_SWISS, index=False, encoding='latin1'):
        """Write Mapinfo format, all geometry types in one file"""

        gdf.crs = WKT_SWISS
        
        # fiona can't write integer columns, convert columns to float
        for col in gdf.columns:
            stype = str(gdf[col].dtype)
            if stype.startswith('int'):
                gdf[col] = gdf[col].astype(float)
                
        # fiona can't write integer columns, convert index to float
        stype = str(gdf.index.dtype)
        if stype.startswith('int'):
            gdf.index = gdf.index.astype(float)
            
        delete_mapinfo_files(tab_name)

        gdf.to_file(tab_name,driver='MapInfo File', index=index, encoding=encoding)    
        return print(len(gdf), 'row(s) written to mapinfo file.')
        
def swiss_wgs(sX,sY):
    """Aprroximation CH1903 -> WGS84 https://de.wikipedia.org/wiki/Schweizer_Landeskoordinaten"""
    sX=str(sX)
    sY=str(sY)

    if sX.strip()[0].isdigit():
        x=float(sX)
        y=float(sY)
        x1 = (x - 600000) / 1000000
        y1 = (y - 200000) / 1000000
        L = 2.6779094 + 4.728982 * x1 + 0.791484 * x1 * y1 + 0.1306 * x1 * y1 * y1 - 0.0436 * x1 * x1 * x1
        p = 16.9023892 + 3.238272 * y1 - 0.270978 * x1 * x1 - 0.002528 * y1 * y1 - 0.0447 * x1 * x1 * y1 - 0.014 * y1 * y1 * y1
        return (L*100/36,p*100/36)
    else:
        return (-1,-1)

def wgs_swiss(sLon, sLat):
    """Aprroximation WGS84 -> CH1903 https://de.wikipedia.org/wiki/Schweizer_Landeskoordinaten"""

    sLon=str(sLon)
    sLat=str(sLat)

    if sLon.strip()[0].isdigit():
        Lon=float(sLon)
        Lat=float(sLat)
        p = (Lat * 3600 - 169028.66) / 10000
        L = (Lon * 3600 - 26782.5) / 10000
        y = 200147.07 + 308807.95 * p + 3745.25 * L * L + 76.63 * p * p - 194.56 * L * L * p + 119.79 * p * p * p
        x = 600072.37 + 211455.93 * L - 10938.51 * L * p - 0.36 * L * p * p - 44.54 * L * L * L
        return (x,y)
    else:
        return (-1,-1)
        
def conv_wgs_swiss (llon, llat, resolution=1):
    lx = []
    ly = []
    for xy in zip(llon, llat):
        x, y = wgs_swiss(*xy)
        x = x // resolution * resolution
        y = y // resolution * resolution
        lx.append(x)
        ly.append(y)
    return lx, ly
    
def run(str_or_list):
    """Better replacement for os.system()"""
    subprocess.run(str_or_list, check=True, shell=True)	

def run_mb(mb_script:str):
    """Run Mapbasic string as mapbasic script
mapinfow.exe and mapbascic : both paths must be set in the PATH env variable!
    """
    wd = os.getcwd()

    path_mb = os.path.join(wd, 'mb.mb')
    path_mbx = os.path.join(wd, 'mb.mbx')
    
    if os.path.isfile(path_mb): os.remove(path_mb)
    if os.path.isfile(path_mbx): os.remove(path_mbx)

    print(path_mb)
    
    with open(path_mb,'w') as fout: 
        fout.write(mb_script)
    
    # Compile
    subprocess.run([r'C:\Program Files (x86)\MapInfo\MapBasic\mapbasic.exe', '-D', path_mb], check=True, shell=True)
    
    # Run
    try:
        subprocess.run([r"C:\Program Files (x86)\MapInfo\Professional\MapInfow.exe", path_mbx, path_mb], check=True, shell=True)
    except subprocess.CalledProcessError as e:
        print(e)

if USE_GEOPANDAS:
    def disagg(vec:gpd.GeoDataFrame):
        """Dissagregate collections and multi geomtries"""
        # Split GeometryCollections
        no_coll = []
        for i, row in vec.iterrows():
            geom = row.geometry
            if geom.type == 'GeometryCollection':
                for part in geom:
                    row2 = row.copy()
                    row2.geometry = part
                    no_coll.append(row2)

            else:
                    no_coll.append(row)           

        # Split Multi geomries
        res = []
        for row in no_coll:
            geom = row.geometry
            if geom.type.startswith('Multi'):
                for part in geom:
                    row2 = row.copy()
                    row2.geometry = part
                    res.append(row2)
            else:
                    res.append(row)

        return gpd.GeoDataFrame(res, crs=vec.crs).reset_index(drop=True)

if USE_GEOPANDAS:
    def write_geojson(vec:gpd.GeoDataFrame, dest):
        """Write only polygons, including attributes"""
        dest = str(dest)

        # WGS 84
        #vec = vec.to_crs({'init': 'epsg:4326'})

        if os.path.isfile(dest):
            os.remove(dest)
            
        vec.to_file(dest, driver='GeoJSON', encoding='utf-8')


def read_loss(los_file):
    los_file = str(los_file)
    basename = os.path.basename(los_file)
    dbf_path = os.path.dirname(los_file)
    dbf_path = os.path.join(dbf_path, 'pathloss.dbf')
    pl = read_dbf(dbf_path)
    pl = pl.set_index('FILE_NAME')
 
    ser = pl.loc[basename]
    ulx = ser.ULXMAP
    uly = ser.ULYMAP
    nx = ser.NCOLS
    ny = ser.NROWS
    res = ser.RESOLUTION
    #print (basename,ulx,uly,nx,ny,res)
    r = np.fromfile(los_file, dtype='int16')
    r.resize(ny, nx)
    df = pd.DataFrame(r)
    df.columns = np.linspace(ulx, ulx+res*nx-res, nx)
    df.index = np.linspace(uly-res*nx, uly-res, ny)
    df = df.sort_index(ascending=False)
    df = df/16
    return df

def show_perc(i:int, iall:int, istep:int):
    if i % istep == 0:
        print(f'{round(100*i/iall,2)}%', end=' ')
          
def write_sqlite(db_file, df_dict):
    db_conn = sqlite3.connect(db_file)
    for tab_name in df_dict:
        df_dict[tab_name].to_sql(tab_name, db_conn, if_exists="replace", index=None)
    db_conn.close()

def write_json(obj, filename, pretty=True):
    if pretty:
        indent = 4
    else:
        indent = None
    with open(filename, 'w', encoding='utf-8') as fout:
        json.dump(obj, fout,  ensure_ascii=False, indent=indent)
        
def days_per_month(first_day, last_day):
    """return dictionary of month with mumber of days"""
    first_day = pd.to_datetime(first_day)
    last_day = pd.to_datetime(last_day)

    l = pd.date_range(first_day.replace(day=1), last_day.replace(day=1), freq='MS')

    dic = {}
    for e in l:
        dic[str(e)[0:7]] = pd.Period(str(e)).daysinmonth
    
    # patch first and last month
    dic[str(l[0])[0:7]] = dic[str(l[0])[0:7]] - first_day.day + 1
    dic[str(l[-1])[0:7]] = last_day.day
    
    return dic
    
def black2trasparent(source_pic, dest_png):
    img = Image.open(source_pic).convert("RGBA")
    data = np.asarray(img).copy()
    data[:, :, 3] = (255 * (data[:, :, :3] != 0).any(axis=2)).astype(np.uint8)
    img2 = Image.fromarray(data)
    img2.save(dest_png, "PNG")

def make_raster_kml(source_file, dest_kml_file, name=None):

    assert pathlib.Path(source_file).is_file(), f'file {source_file} not found.'
    # creade destination filenames
    dest_kml = pathlib.Path(dest_kml_file).with_suffix('.kml')
    dest_png = pathlib.Path(dest_kml_file).with_suffix('.png')

    print('create tmp.tif as WGS84')
    cmd = f'gdalwarp -s_srs EPSG:21781 -t_srs EPSG:4326 -overwrite "{source_file}" tmp.tif'
    os.system(cmd)
    
    print(f'convert to transparent {dest_png}')
    black2trasparent('tmp.tif', dest_png)
    
    ds = rasterio.open('tmp.tif')
    x0, y0, x1, y1 = ds.bounds
    print('bounds:', x0, x1, y0, y1)
    
    if name is None:
        name = dest_kml.stem

    s = f"""<kml>
        <GroundOverlay>
          <name>{name}</name>
          <Icon>
            <href>{dest_png.name}</href>
          </Icon>
          <LatLonBox>
            <west>{x0}</west>
            <east>{x1}</east>
            <south>{y0}</south>
            <north>{y1}</north>
          </LatLonBox>
        </GroundOverlay>
    </kml>
    """
    print('write', dest_kml)
    dest_kml.write_text(s)
    
def beep():
    import winsound
    winsound.PlaySound(r"C:\Windows\Media\tada.wav", winsound.SND_ASYNC)
    

def write_excel_with_template(df, target, template, sheet_name, header_start='A2'):
    """ target must be full windows name"""
    # open workbook get sheet
    wb = xw.Book(template)
    sheet = wb.sheets[sheet_name]
    
    # # check if datatframe headers equal to template headers
    lh = sheet.range(header_start).expand('right').value
    for i, col in enumerate(df.columns):
        assert col == lh[i], f'header in column {i} "{col}" != "{lh[i]}"'
        
    # write the datafram
    sheet.range(header_start).options(index=False).value = df
    
    # switch off wrap text
    full_range = sheet.range('A2').expand()
    sheet.range(full_range).api.WrapText = False
    
    # save excel
    wb.save(target)
    
    # close excel
    wb.app.quit()
    
    
class Excel:
    def __init__(self, dest_file):
        self.writer = pd.ExcelWriter(dest_file, engine='xlsxwriter')
        
    def write_sheet(self, df, sheet_name, header_font_color='#000000', header_bg_color='#FFFFFF', max_col_len=50):
        assert not df.columns.duplicated().any(), 'Error: duplicated columns'
        df.to_excel(excel_writer=self.writer, sheet_name=sheet_name, index=None)
        
        workbook  = self.writer.book
        worksheet = self.writer.sheets[sheet_name]
        rows = len(df)
        cols = len(df.columns)
        
        # set header format
        header_format = workbook.add_format({'font_color': header_font_color, 'bg_color': header_bg_color, 'bold': True, 'align': 'left'}) 
        for i, s in enumerate(df.columns):
            worksheet.write(0, i, s, header_format)
            
        # autofilter
        worksheet.autofilter(0, 0, rows, cols -1)
        
        # autofit
        for idx, col in enumerate(df):  # loop through all columns
            series = df[col]
            max_len = max((
                series.astype(str).map(len).max(),  # len of largest item
                len(str(series.name))  # len of column name/header
                )) + 1  # adding a little extra space
            if max_len > max_col_len:
                max_len = max_col_len
            worksheet.set_column(idx, idx, max_len)  # set column width

        # freeze first row
        worksheet.freeze_panes(1, 0)
        
    def close(self):
        self.writer.close()

    def save(self):
        self.writer.close()
        
     
def datetime2date(df):
    df = df.copy()
    dtypes = df.dtypes
    for col_name, dtype in dtypes.items():
        if str(dtype).startswith('datetime'):
            df[col_name] = pd.to_datetime(df[col_name]).dt.date
    return df


def stylize(source, dest, template):
    print(f'loading workbook "{source}"')
    wb_nostyle = openpyxl.load_workbook(source)
    wb_template = openpyxl.load_workbook(template)

    # create explanation row above header
    for sheetname in wb_nostyle.sheetnames:
        print(f"loading sheet '{sheetname}'")
        ws_nostyle = wb_nostyle[sheetname]
        ws_template = wb_template[sheetname]


        ws_nostyle.insert_rows(1)
        for col_index, cell in enumerate(ws_template[1], start=1):
            new_cell = ws_nostyle.cell(row=1, column=col_index, value=cell.value)

            # Copy formatting
            new_cell.font = copy(cell.font)
            new_cell.fill = copy(cell.fill)
            new_cell.border = copy(cell.border)
            new_cell.alignment = copy(cell.alignment)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)


            col_letter = cell.column_letter
            ws_nostyle.column_dimensions[col_letter].width = ws_template.column_dimensions[col_letter].width


        # set row heigt of explanation row
        ws_nostyle.row_dimensions[1].height = ws_template.row_dimensions[1].height

        # Set format of header
        for col_index, cell in enumerate(ws_template[2], start=1):

            # check if header values are correct
            value_nostyle = ws_nostyle.cell(row=2, column=col_index).value
            value_template = cell.value
            letter = openpyxl.utils.get_column_letter(col_index)
            assert value_nostyle == value_template, f'ERROR in Column {letter} {value_nostyle} != {value_template}'

            new_cell = ws_nostyle.cell(row=2, column=col_index, value=cell.value)


            # Copy formatting
            new_cell.font = copy(cell.font)
            new_cell.fill = copy(cell.fill)
            new_cell.border = copy(cell.border)
            new_cell.alignment = copy(cell.alignment)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)


        # Determine the range of the table
        max_row = ws_nostyle.max_row
        max_col = ws_nostyle.max_column
        end_col_letter = openpyxl.utils.get_column_letter(max_col)

        # Set AutoFilter for the whole table
        full_range = f"A2:{end_col_letter}{max_row}"
        ws_nostyle.auto_filter.ref = full_range


        # switching off wordwrap
        print('switching off wordwrap', end=' ')
        no_wrap_alignment = openpyxl.styles.Alignment(wrap_text=False)
        for i, row in enumerate(ws_nostyle[full_range]):
            if i % 1000 == 0:
                perc = round(100*i/max_row)
                # print(f'{perc}%', end=' ')
                # print('.', end=' ')
            for cell in row:
                cell.alignment = no_wrap_alignment
        print()

    print(f'saving "{dest}"')
    wb_nostyle.save(dest)
